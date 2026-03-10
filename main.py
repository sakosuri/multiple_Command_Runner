#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""

Cisco Catalyst Center Bulk Command Runner

Reads multiple CLI commands from an Excel file and executes them
on all devices managed by Catalyst Center (formerly known as DNA Center).

Based on the original Cisco Sample Code by Gabriel Zapodeanu.
"""

import requests
import json
import urllib3
import time
import sys
import os
from datetime import datetime

from openpyxl import load_workbook, Workbook
from urllib3.exceptions import InsecureRequestWarning
from requests.auth import HTTPBasicAuth

from config import DNAC_URL, DNAC_PASS, DNAC_USER

urllib3.disable_warnings(InsecureRequestWarning)

DNAC_AUTH = HTTPBasicAuth(DNAC_USER, DNAC_PASS)

# ─── Configuration ─────────────────────────────────────────────
INPUT_EXCEL = 'commands.xlsx'        # Input Excel file with commands
INPUT_SHEET = 'Commands'             # Sheet name in input Excel
OUTPUT_EXCEL = 'command_output.xlsx' # Output Excel file for results
COMMAND_TIMEOUT = 0                  # Timeout for command runner API
TASK_POLL_INTERVAL = 2               # Seconds between task status polls
MAX_TASK_WAIT = 60                   # Maximum seconds to wait for a task
# ────────────────────────────────────────────────────────────────


def pprint(json_data):
    """Pretty print JSON formatted data."""
    print(json.dumps(json_data, indent=4, separators=(' , ', ' : ')))


def get_dnac_jwt_token(dnac_auth):
    """
    Create the authorization token required to access Cisco Catalyst Center.
    :param dnac_auth: Cisco Catalyst Center Basic Auth string
    :return: Cisco Catalyst Center JWT Token
    """
    url = DNAC_URL + '/dna/system/api/v1/auth/token'
    header = {'content-type': 'application/json'}
    response = requests.post(url, auth=dnac_auth, headers=header, verify=False)
    response.raise_for_status()
    dnac_jwt_token = response.json()['Token']
    return dnac_jwt_token


def get_all_device_info(dnac_jwt_token):
    """
    Return all network devices info, handling pagination.
    :param dnac_jwt_token: Cisco Catalyst Center token
    :return: list of all device dictionaries
    """
    all_devices = []
    offset = 1
    limit = 500  # max per page

    header = {'content-type': 'application/json', 'x-auth-token': dnac_jwt_token}

    while True:
        url = (DNAC_URL +
               f'/dna/intent/api/v1/network-device?offset={offset}&limit={limit}')
        response = requests.get(url, headers=header, verify=False)
        response.raise_for_status()
        devices = response.json().get('response', [])
        if not devices:
            break
        all_devices.extend(devices)
        if len(devices) < limit:
            break
        offset += limit

    return all_devices


def get_legit_cli_command_runner(dnac_jwt_token):
    """
    Get all legit CLI command keywords supported by the command runner API.
    :param dnac_jwt_token: Cisco Catalyst Center token
    :return: list of supported CLI command keywords
    """
    url = DNAC_URL + '/dna/intent/api/v1/network-device-poller/cli/legit-reads'
    header = {'content-type': 'application/json', 'x-auth-token': dnac_jwt_token}
    response = requests.get(url, headers=header, verify=False)
    response.raise_for_status()
    return response.json()['response']


def get_content_file_id(file_id, dnac_jwt_token):
    """
    Download the file specified by the file_id.
    :param file_id: file id
    :param dnac_jwt_token: Cisco Catalyst Center token
    :return: file content as JSON
    """
    url = DNAC_URL + '/dna/intent/api/v1/file/' + file_id
    header = {'content-type': 'application/json', 'x-auth-token': dnac_jwt_token}
    response = requests.get(url, headers=header, verify=False, stream=True)
    response.raise_for_status()
    return response.json()


def check_task_id_output(task_id, dnac_jwt_token):
    """
    Poll the task until it completes or times out.
    :param task_id: task id
    :param dnac_jwt_token: Cisco Catalyst Center token
    :return: task output dict
    """
    url = DNAC_URL + '/dna/intent/api/v1/task/' + task_id
    header = {'content-type': 'application/json', 'x-auth-token': dnac_jwt_token}

    elapsed = 0
    while elapsed < MAX_TASK_WAIT:
        try:
            task_response = requests.get(url, headers=header, verify=False)
            task_response.raise_for_status()
            task_output = task_response.json()['response']

            # Check for failure
            if task_output.get('isError'):
                return task_output

            # Check if progress contains fileId (indicates completion)
            progress = task_output.get('progress', '')
            file_info = json.loads(progress)
            if 'fileId' in file_info:
                return task_output
        except (json.JSONDecodeError, KeyError, TypeError):
            pass

        time.sleep(TASK_POLL_INTERVAL)
        elapsed += TASK_POLL_INTERVAL

    raise TimeoutError(
        f"Task {task_id} did not complete within {MAX_TASK_WAIT} seconds."
    )


def execute_command_on_device(command, device_id, device_name, dnac_jwt_token):
    """
    Execute a single CLI command on a single device and return the output.
    :param command: CLI command string
    :param device_id: Catalyst Center device UUID
    :param device_name: device hostname (for logging)
    :param dnac_jwt_token: Cisco Catalyst Center token
    :return: tuple (status, output_text)
             status  = 'SUCCESS' | 'FAILURE' | 'BLACKLISTED' | 'ERROR'
    """
    payload = {
        "commands": [command],
        "deviceUuids": [device_id],
        "timeout": COMMAND_TIMEOUT
    }
    url = DNAC_URL + '/dna/intent/api/v1/network-device-poller/cli/read-request'
    header = {'content-type': 'application/json', 'x-auth-token': dnac_jwt_token}

    try:
        response = requests.post(
            url, data=json.dumps(payload), headers=header, verify=False
        )
        response.raise_for_status()
        response_json = response.json()

        # Check for immediate errors
        if 'response' not in response_json:
            return ('ERROR', f"Unexpected API response: {response_json}")

        if 'taskId' not in response_json['response']:
            detail = response_json['response'].get('detail', 'Unknown error')
            return ('ERROR', detail)

        task_id = response_json['response']['taskId']

        # Wait for task completion
        time.sleep(TASK_POLL_INTERVAL)
        task_result = check_task_id_output(task_id, dnac_jwt_token)

        # Check for task-level error
        if task_result.get('isError'):
            return ('ERROR', task_result.get('failureReason', 'Task failed'))

        file_info = json.loads(task_result['progress'])
        file_id = file_info['fileId']

        # Retrieve file output
        time.sleep(TASK_POLL_INTERVAL)
        file_output = get_content_file_id(file_id, dnac_jwt_token)

        command_responses = file_output[0]['commandResponses']

        if command_responses.get('SUCCESS') and command in command_responses['SUCCESS']:
            return ('SUCCESS', command_responses['SUCCESS'][command])
        elif command_responses.get('FAILURE') and command in command_responses['FAILURE']:
            return ('FAILURE', command_responses['FAILURE'][command])
        elif (command_responses.get('BLACKLISTED')
              and command in command_responses['BLACKLISTED']):
            return ('BLACKLISTED', command_responses['BLACKLISTED'][command])
        else:
            return ('ERROR', f"Command not found in response: {command_responses}")

    except TimeoutError as e:
        return ('ERROR', str(e))
    except Exception as e:
        return ('ERROR', f"Exception: {str(e)}")


def read_commands_from_excel(file_path, sheet_name):
    """
    Read CLI commands from column A of the specified Excel sheet.
    :param file_path: path to the Excel file
    :param sheet_name: name of the sheet
    :return: list of command strings
    """
    if not os.path.exists(file_path):
        print(f"\nERROR: Input file '{file_path}' not found.")
        sys.exit(1)

    wb = load_workbook(file_path, read_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"\nERROR: Sheet '{sheet_name}' not found in '{file_path}'.")
        print(f"Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]
    commands = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        cell_value = row[0]
        if cell_value is not None:
            cmd = str(cell_value).strip()
            if cmd:
                commands.append(cmd)
    wb.close()

    if not commands:
        print(f"\nERROR: No commands found in '{file_path}' sheet '{sheet_name}'.")
        sys.exit(1)

    return commands


def write_results_to_excel(results, file_path):
    """
    Write command execution results to an Excel file.
    :param results: list of dicts with keys:
                    'Device', 'Device ID', 'Command', 'Status', 'Output'
    :param file_path: output file path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Command Output'

    # Header row
    headers = ['Device Hostname', 'Device ID', 'Command', 'Status', 'Output']
    ws.append(headers)

    # Bold header
    from openpyxl.styles import Font
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    # Data rows
    for r in results:
        ws.append([
            r['Device'],
            r['Device ID'],
            r['Command'],
            r['Status'],
            r['Output']
        ])

    # Auto-adjust column widths (approximate)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    # Limit scan to first 200 chars to avoid slow processing
                    max_length = max(max_length, min(len(str(cell.value)), 80))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

    wb.save(file_path)
    print(f"\n✅ Results saved to '{file_path}'")


def main():
    """
    Main workflow:
    1. Authenticate with Catalyst Center
    2. Read commands from Excel
    3. Retrieve all managed devices
    4. Validate commands against supported keyword list
    5. Execute each command on each device
    6. Collect and save results to an output Excel file
    """

    start_time = datetime.now()
    print('=' * 70)
    print('  Cisco Catalyst Center — Bulk Command Runner')
    print(f'  Started at: {start_time.strftime("%Y-%m-%d %H:%M:%S")}')
    print('=' * 70)

    # ── Step 1: Authenticate ───────────────────────────────────
    print('\n[1/5] Authenticating with Catalyst Center...')
    dnac_token = get_dnac_jwt_token(DNAC_AUTH)
    print('      Authentication successful.')

    # ── Step 2: Read commands from Excel ───────────────────────
    print(f'\n[2/5] Reading commands from "{INPUT_EXCEL}" '
          f'(sheet: "{INPUT_SHEET}")...')
    commands = read_commands_from_excel(INPUT_EXCEL, INPUT_SHEET)
    print(f'      Found {len(commands)} command(s):')
    for i, cmd in enumerate(commands, 1):
        print(f'        {i}. {cmd}')

    # ── Step 3: Retrieve all devices ──────────────────────────
    print('\n[3/5] Retrieving all managed devices...')
    all_devices = get_all_device_info(dnac_token)

    # Filter to only reachable devices with a hostname
    managed_devices = [
        d for d in all_devices
        if d.get('hostname') and d.get('reachabilityStatus') == 'Reachable'
    ]
    print(f'      Total devices in inventory : {len(all_devices)}')
    print(f'      Reachable devices           : {len(managed_devices)}')

    if not managed_devices:
        print('\nERROR: No reachable devices found. Exiting.')
        sys.exit(1)

    print('\n      Device list:')
    for i, d in enumerate(managed_devices, 1):
        print(f'        {i:>3}. {d["hostname"]:<40} '
              f'({d.get("managementIpAddress", "N/A")})')

    # ── Step 4: Validate commands ─────────────────────────────
    print('\n[4/5] Validating commands against supported keyword list...')
    supported_keywords = get_legit_cli_command_runner(dnac_token)

    valid_commands = []
    for cmd in commands:
        keyword = cmd.split(' ')[0]
        if keyword in supported_keywords:
            valid_commands.append(cmd)
            print(f'        ✔ "{cmd}" — supported')
        else:
            print(f'        ✘ "{cmd}" — NOT supported (keyword "{keyword}" '
                  f'not in allowed list). Skipping.')

    if not valid_commands:
        print('\nERROR: None of the provided commands are supported. Exiting.')
        sys.exit(1)

    # ── Step 5: Execute commands on all devices ───────────────
    total_tasks = len(valid_commands) * len(managed_devices)
    print(f'\n[5/5] Executing {len(valid_commands)} command(s) on '
          f'{len(managed_devices)} device(s) '
          f'({total_tasks} total executions)...\n')

    results = []
    counter = 0

    for device in managed_devices:
        device_name = device['hostname']
        device_id = device['id']
        device_ip = device.get('managementIpAddress', 'N/A')

        print(f'  ── Device: {device_name} ({device_ip}) ──')

        for cmd in valid_commands:
            counter += 1
            print(f'      [{counter}/{total_tasks}] Running: "{cmd}" ... ',
                  end='', flush=True)

            status, output = execute_command_on_device(
                cmd, device_id, device_name, dnac_token
            )

            print(f'{status}')

            results.append({
                'Device': device_name,
                'Device ID': device_id,
                'Command': cmd,
                'Status': status,
                'Output': output
            })

    # ── Save results ──────────────────────────────────────────
    write_results_to_excel(results, OUTPUT_EXCEL)

    # ── Summary ───────────────────────────────────────────────
    end_time = datetime.now()
    duration = end_time - start_time

    success_count = sum(1 for r in results if r['Status'] == 'SUCCESS')
    failure_count = sum(1 for r in results if r['Status'] != 'SUCCESS')

    print('\n' + '=' * 70)
    print('  Execution Summary')
    print('=' * 70)
    print(f'  Total executions : {len(results)}')
    print(f'  Successful       : {success_count}')
    print(f'  Failed/Other     : {failure_count}')
    print(f'  Duration         : {duration}')
    print(f'  Output file      : {OUTPUT_EXCEL}')
    print('=' * 70 + '\n')


if __name__ == "__main__":
    sys.exit(main())
